#!/usr/bin/env python3
"""
EDPA Engine — Evidence-Driven Proportional Allocation

Standalone Python implementation of the EDPA calculation engine.
Computes derived hours from delivery evidence stored in .edpa/backlog/.

Usage:
    python .claude/edpa/scripts/engine.py --edpa-root .edpa --iteration PI-2026-1.3
    python .claude/edpa/scripts/engine.py --edpa-root .edpa --iteration PI-2026-1.3 --mode full
    python .claude/edpa/scripts/engine.py --demo  # Run with built-in sample data

    # Legacy mode (requires external item gathering):
    python .claude/edpa/scripts/engine.py --capacity cap.yaml --heuristics h.yaml --iteration PI-2026-1.3
"""

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    import yaml
except ImportError:
    print("ERROR: PyYAML required. Install with: pip install pyyaml")
    sys.exit(1)


def get_version():
    """Read version from plugin.json (single source of truth)."""
    for candidate in [
        Path(__file__).parent.parent.parent / ".claude-plugin" / "plugin.json",
        Path(__file__).parent.parent.parent.parent / ".claude" / ".claude-plugin" / "plugin.json",
    ]:
        try:
            if candidate.exists():
                with open(candidate) as f:
                    return json.load(f).get("version", "unknown")
        except Exception:
            continue
    return "unknown"


VERSION = get_version()


def load_yaml(path):
    with open(path) as f:
        return yaml.safe_load(f)


def gh_json(cmd):
    """Run gh CLI command and parse JSON output."""
    try:
        result = subprocess.run(
            ["gh"] + cmd.split() + ["--json", "number,title,assignees,labels,body"],
            capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            return json.loads(result.stdout)
    except (subprocess.TimeoutExpired, FileNotFoundError, json.JSONDecodeError):
        pass
    return []


def extract_item_refs(text):
    """Extract work item references (S-123, F-45, E-7) from text."""
    if not text:
        return []
    return re.findall(r'[SFEITD]-\d+', text)


def detect_evidence(people, items, iteration_id):
    """
    Detect contribution evidence from GitHub data.

    Returns: dict of {(person_id, item_id): {"signals": [...], "evidence_score": float, "cw": float}}
    """
    evidence = {}

    for item in items:
        item_id = item.get("id", "")
        assignees = [a.get("login", "") for a in item.get("assignees", [])]

        for person in people:
            pid = person["id"]

            # Check evidence_scope
            scope = person.get("evidence_scope")
            if scope:
                import fnmatch
                if not any(fnmatch.fnmatch(item_id, pattern) for pattern in scope):
                    # Item doesn't match this contract's scope
                    if not person.get("evidence_default", False):
                        continue  # Skip — not in scope and not default

            signals = []
            score = 0.0

            # Check assignee
            if pid in assignees or person.get("email", "") in assignees:
                score += 4.0
                signals.append("assignee")

            # Check /contribute commands in body
            body = item.get("body", "") or ""
            contribute_pattern = rf'/contribute\s+@{re.escape(pid)}\s+weight:([0-9.]+)'
            contribute_match = re.search(contribute_pattern, body)
            if contribute_match:
                score += 3.0
                signals.append("contribute_command")

            # Check PR author (simplified — looks at linked PRs)
            if item.get("pr_author") == pid:
                score += 2.0
                signals.append("pr_author")

            # Check commit author
            if item.get("commit_authors") and pid in item["commit_authors"]:
                score += 1.0
                signals.append("commit_author")

            # Check PR reviewer
            if item.get("pr_reviewers") and pid in item["pr_reviewers"]:
                score += 1.0
                signals.append("pr_reviewer")

            # Check comments
            if item.get("commenters") and pid in item["commenters"]:
                score += 0.5
                signals.append("issue_comment")

            if signals:
                evidence[(pid, item_id)] = {
                    "signals": signals,
                    "evidence_score": score,
                    "manual_cw": float(contribute_match.group(1)) if contribute_match else None
                }

    return evidence


def compute_cw(evidence_entry, heuristics, person_role=None):
    """Compute Contribution Weight from evidence signals.

    Uses role_overrides (Monte Carlo calibrated) when person_role is known,
    falling back to generic role_weights otherwise.
    """
    if evidence_entry.get("manual_cw") is not None:
        return evidence_entry["manual_cw"]

    signal_to_role = {
        "assignee": "owner",
        "contribute_command": "key",
        "pr_author": "key",
        "commit_author": "reviewer",
        "pr_reviewer": "reviewer",
        "issue_comment": "consulted",
    }

    role_priority = ["assignee", "contribute_command", "pr_author",
                     "commit_author", "pr_reviewer", "issue_comment"]

    role_weights = heuristics.get("role_weights", {})
    role_overrides = heuristics.get("role_overrides", {})

    for signal in role_priority:
        if signal in evidence_entry["signals"]:
            evidence_role = signal_to_role[signal]

            # Check role_overrides first (Monte Carlo calibrated)
            if person_role and person_role in role_overrides:
                override = role_overrides[person_role]
                if evidence_role in override:
                    return override[evidence_role]

            # Fallback to generic weights
            return role_weights.get(evidence_role, 0.15)

    return 0.15


def run_edpa(capacity_config, heuristics, items, mode="simple"):
    """
    Run the core EDPA calculation.

    Returns: list of person results with derived hours.
    """
    people = capacity_config.get("people", [])
    threshold = heuristics.get("evidence_threshold", 1.0)
    iteration_id = "computed"

    # Detect evidence
    evidence = detect_evidence(people, items, iteration_id)

    results = []

    for person in people:
        pid = person["id"]
        cpi = person.get("capacity_per_iteration")
        capacity = cpi if cpi is not None else person.get("capacity", 0)
        person_items = []

        for item in items:
            item_id = item["id"]
            key = (pid, item_id)

            if key not in evidence:
                continue

            ev = evidence[key]
            if ev["evidence_score"] < threshold:
                continue

            cw = compute_cw(ev, heuristics, person_role=person.get("role"))
            js = item.get("job_size", 0)

            if js <= 0:
                continue

            if mode == "full":
                # Compute Relevance Signal
                max_es = max(
                    (evidence.get((p["id"], item_id), {}).get("evidence_score", 0)
                     for p in people),
                    default=1.0
                )
                rs = min(ev["evidence_score"] / max_es, 1.0) if max_es > 0 else 1.0
            else:
                rs = 1.0

            score = js * cw * rs

            person_items.append({
                "id": item_id,
                "level": item.get("level", "Story"),
                "js": js,
                "cw": round(cw, 4),
                "rs": round(rs, 4),
                "score": round(score, 4),
                "evidence": ev["signals"],
            })

        # Calculate derived hours
        sum_scores = sum(pi["score"] for pi in person_items)

        for pi in person_items:
            if sum_scores > 0:
                ratio = pi["score"] / sum_scores
                hours = ratio * capacity
            else:
                ratio = 0.0
                hours = 0.0
            pi["ratio"] = round(ratio, 6)
            pi["hours"] = round(hours, 2)

        # Normalize: adjust last item so sum exactly equals capacity
        if person_items and sum_scores > 0:
            rounded_sum = sum(pi["hours"] for pi in person_items)
            diff = round(capacity - rounded_sum, 2)
            if diff != 0:
                person_items[-1]["hours"] = round(person_items[-1]["hours"] + diff, 2)

        total_derived = sum(pi["hours"] for pi in person_items)

        # Validate invariants
        invariant_ok = True
        if person_items:
            if abs(total_derived - capacity) > 0.1:
                invariant_ok = False
            ratio_sum = sum(pi["ratio"] for pi in person_items)
            if abs(ratio_sum - 1.0) > 0.001:
                invariant_ok = False
            if any(pi["hours"] < 0 for pi in person_items):
                invariant_ok = False

        results.append({
            "id": pid,
            "name": person.get("name", pid),
            "role": person.get("role", ""),
            "capacity": capacity,
            "total_derived": round(total_derived, 2),
            "items": person_items,
            "invariant_ok": invariant_ok,
        })

    return results


def load_heuristics(edpa_root):
    """Load CW heuristics from .edpa/config/, with fallback chain.

    Tries: heuristics.yaml → cw_heuristics.yaml → template in .claude/
    """
    edpa_root = Path(edpa_root)
    for name in ("heuristics.yaml", "cw_heuristics.yaml"):
        path = edpa_root / "config" / name
        if path.exists():
            return load_yaml(path)
    # Fallback to template (installed plugin)
    template = edpa_root.parent / ".claude" / "edpa" / "templates" / "cw_heuristics.yaml.tmpl"
    if template.exists():
        return load_yaml(template)
    return {"evidence_threshold": 1.0, "role_weights": {"owner": 1.0, "key": 0.6, "reviewer": 0.25, "consulted": 0.15}}


def load_backlog_items(edpa_root, iteration_id=None):
    """Read .edpa/backlog/ YAML files and convert to engine item format.

    Each backlog YAML has: id, type, title, js, status, assignee, contributors
    Engine expects: id, level, job_size, assignees, body, pr_author, commit_authors, pr_reviewers, commenters

    Args:
        edpa_root: Path to .edpa/ directory
        iteration_id: If given, only include items matching this iteration. If None, include all Done items.

    Returns:
        List of item dicts in engine format, plus a dict of manual CW overrides.
    """
    edpa_root = Path(edpa_root)
    backlog_dir = edpa_root / "backlog"
    if not backlog_dir.exists():
        return [], {}

    items = []
    manual_cw_overrides = {}  # {(person_id, item_id): cw_value}

    type_dirs = {
        "stories": "Story",
        "features": "Feature",
        "epics": "Epic",
        "initiatives": "Initiative",
        "defects": "Defect",
    }

    for dir_name, level in type_dirs.items():
        type_dir = backlog_dir / dir_name
        if not type_dir.exists():
            continue

        for yaml_file in sorted(type_dir.glob("*.yaml")):
            try:
                data = load_yaml(yaml_file)
            except Exception:
                continue

            if not data or not isinstance(data, dict):
                continue

            item_id = data.get("id", yaml_file.stem)
            status = data.get("status", "")

            # Filter: only Done items
            if status.lower() not in ("done", "closed", "accepted"):
                continue

            # Filter by iteration — SAFe hierarchy-aware:
            #   Story → exact iteration match (e.g., PI-2026-1.1)
            #   Feature → PI match (e.g., PI-2026-1 matches PI-2026-1.x)
            #   Epic/Initiative → always included if Done (cross-PI)
            item_type = data.get("type", level)
            item_iter = data.get("iteration", "")

            if iteration_id:
                if item_type == "Story":
                    if item_iter != iteration_id:
                        continue
                elif item_type == "Feature":
                    pi_prefix = iteration_id.rsplit(".", 1)[0]
                    if item_iter != pi_prefix and item_iter != iteration_id:
                        continue
                # Epic + Initiative: always included if Done

            js = data.get("js") or data.get("job_size", 0)
            if not js or js <= 0:
                continue

            # Map contributors to engine evidence fields
            assignees = []
            pr_author = None
            commit_authors = []
            pr_reviewers = []
            commenters = []
            body_parts = []
            contributors = data.get("contributors", []) or []

            # Assignee from top-level field
            assignee = data.get("assignee") or data.get("owner")
            if assignee:
                assignees.append({"login": assignee})

            for contrib in contributors:
                if not isinstance(contrib, dict):
                    continue
                person = contrib.get("person", "")
                role = (contrib.get("role", "") or "").lower()
                cw = contrib.get("cw")

                # Store manual CW override if present
                if cw is not None:
                    manual_cw_overrides[(person, item_id)] = float(cw)

                # Map contributor role to engine evidence fields
                if role == "owner":
                    if not any(a.get("login") == person for a in assignees):
                        assignees.append({"login": person})
                elif role == "key":
                    if pr_author is None:
                        pr_author = person
                    commit_authors.append(person)
                elif role == "reviewer":
                    pr_reviewers.append(person)
                elif role == "consulted":
                    commenters.append(person)

                # Generate /contribute command for manual CW
                if cw is not None:
                    body_parts.append(f"/contribute @{person} weight:{cw}")

            items.append({
                "id": item_id,
                "level": data.get("type", level),
                "job_size": js,
                "assignees": assignees,
                "body": "\n".join(body_parts) if body_parts else "",
                "pr_author": pr_author,
                "commit_authors": commit_authors,
                "pr_reviewers": pr_reviewers,
                "commenters": commenters,
            })

    return items, manual_cw_overrides


def generate_demo_data():
    """Generate sample data for demonstration (multi-contract).

    Alice is split into two contracts:
      - alice-arch  (Arch, 40h) — scoped to Stories (S-*), evidence_default=true
      - alice-pm    (PM,  20h) — scoped to Epics/Features (E-*, F-*)
    Total team capacity: 40 + 20 + 80 + 60 = 200h.
    """
    capacity = {
        "teams": [
            {"id": "Alpha", "planning_factor": 0.8},
        ],
        "people": [
            {"id": "alice-arch", "name": "Alice (Arch)", "role": "Arch", "team": "Alpha",
             "fte": 0.5, "capacity_per_iteration": 40, "email": "alice@example.com",
             "evidence_scope": ["S-*"], "evidence_default": True},
            {"id": "alice-pm", "name": "Alice (PM)", "role": "PM", "team": "Alpha",
             "fte": 0.25, "capacity_per_iteration": 20, "email": "alice@example.com",
             "evidence_scope": ["E-*", "F-*"]},
            {"id": "bob", "name": "Bob (Dev)", "role": "Dev", "team": "Alpha",
             "fte": 1.0, "capacity_per_iteration": 80, "email": "bob@example.com"},
            {"id": "carol", "name": "Carol (Dev)", "role": "Dev", "team": "Alpha",
             "fte": 0.75, "capacity_per_iteration": 60, "email": "carol@example.com"},
        ]
    }

    heuristics = {
        "version": "2.2",
        "evidence_threshold": 1.0,
        "role_weights": {"owner": 1.0, "key": 0.6, "reviewer": 0.25, "consulted": 0.15},
        "role_overrides": {
            "BO":   {"owner": 1.00, "key": 0.60, "reviewer": 0.35, "consulted": 0.30},
            "PM":   {"owner": 1.00, "key": 0.60, "reviewer": 0.25, "consulted": 0.20},
            "Arch": {"owner": 1.00, "key": 0.60, "reviewer": 0.30, "consulted": 0.15},
            "Dev":  {"owner": 1.00, "key": 0.60, "reviewer": 0.25, "consulted": 0.15},
        },
        "signals": {"assignee": 4.0, "contribute_command": 3.0, "pr_author": 2.0,
                     "commit_author": 1.0, "pr_reviewer": 1.0, "issue_comment": 0.5},
    }

    items = [
        {"id": "S-101", "level": "Story", "job_size": 5,
         "assignees": [{"login": "bob"}],
         "body": "", "pr_author": "bob", "commit_authors": ["bob", "carol"],
         "pr_reviewers": ["alice-arch"], "commenters": []},
        {"id": "S-102", "level": "Story", "job_size": 8,
         "assignees": [{"login": "carol"}],
         "body": "/contribute @alice-arch weight:0.6", "pr_author": "carol",
         "commit_authors": ["carol"], "pr_reviewers": ["bob"],
         "commenters": ["alice-arch"]},
        {"id": "S-103", "level": "Story", "job_size": 3,
         "assignees": [{"login": "bob"}],
         "body": "", "pr_author": "bob", "commit_authors": ["bob"],
         "pr_reviewers": ["alice-arch"], "commenters": []},
        {"id": "F-10", "level": "Feature", "job_size": 13,
         "assignees": [{"login": "alice-pm"}],
         "body": "", "pr_author": None, "commit_authors": [],
         "pr_reviewers": [], "commenters": ["bob", "carol"]},
        {"id": "S-104", "level": "Story", "job_size": 5,
         "assignees": [{"login": "carol"}],
         "body": "", "pr_author": "carol", "commit_authors": ["carol", "bob"],
         "pr_reviewers": ["alice-arch"], "commenters": []},
        {"id": "E-10", "level": "Epic", "job_size": 21,
         "assignees": [{"login": "alice-pm"}],
         "body": "", "pr_author": None, "commit_authors": [],
         "pr_reviewers": [], "commenters": ["bob"]},
    ]

    return capacity, heuristics, items


def print_summary(results, mode, iteration_id, planning_factor=0.8):
    """Print human-readable summary table."""
    print(f"\n{'='*70}")
    print(f"EDPA {VERSION} — Iteration {iteration_id} ({mode} mode)")
    print(f"{'='*70}")
    print(f"{'Person':<25} {'Role':<8} {'Capacity':>8} {'Derived':>8} {'Items':>6} {'OK':>4}")
    print(f"{'-'*70}")

    team_capacity = 0
    team_derived = 0
    all_ok = True

    for r in results:
        ok = "OK" if r["invariant_ok"] else "FAIL"
        if not r["invariant_ok"]:
            all_ok = False
        team_capacity += r["capacity"]
        team_derived += r["total_derived"]
        print(f"{r['name']:<25} {r['role']:<8} {r['capacity']:>7}h {r['total_derived']:>7}h {len(r['items']):>6} {ok:>4}")

    print(f"{'-'*70}")
    team_planning = round(team_capacity * planning_factor, 1)
    print(f"{'TEAM TOTAL':<25} {'':8} {team_capacity:>7}h {team_derived:>7}h")
    print(f"{'PLANNING CAPACITY':<25} {'':8} {team_planning:>7}h  (factor: {planning_factor})")
    print(f"\nAll invariants passed: {'YES' if all_ok else 'NO'}")

    # Per-person detail
    for r in results:
        if r["items"]:
            print(f"\n--- {r['name']} ({r['capacity']}h) ---")
            print(f"  {'Item':<10} {'Level':<8} {'JS':>4} {'CW':>6} {'Score':>7} {'Ratio':>7} {'Hours':>7}")
            for item in r["items"]:
                print(f"  {item['id']:<10} {item['level']:<8} {item['js']:>4} {item['cw']:>6.2f} {item['score']:>7.2f} {item['ratio']:>6.1%} {item['hours']:>6.1f}h")


def show_status(edpa_root):
    """Show EDPA setup status — config, team, iterations."""
    print(f"EDPA {VERSION} — Status")
    print("=" * 40)

    if not edpa_root.exists():
        print(f"\n✗ .edpa/ not found at {edpa_root}")
        print("  Run: /edpa setup \"Project Name\"")
        return

    print(f"✓ .edpa/ found at {edpa_root}")

    # People config
    people_path = edpa_root / "config" / "people.yaml"
    if people_path.exists():
        people = load_yaml(people_path) or {}
        team = people.get("people", [])
        total_fte = sum(p.get("fte", 0) for p in team)
        total_cap = sum(p.get("capacity_per_iteration", p.get("capacity", 0)) for p in team)
        print(f"✓ people.yaml — {len(team)} members, {total_fte:.1f} FTE, {total_cap:.0f}h/iteration")
        for p in team:
            cap = p.get("capacity_per_iteration", p.get("capacity", 0))
            print(f"    {p.get('name', p.get('id', '?')):<25} {p.get('role', '?'):<8} {p.get('fte', 0):.1f} FTE  {cap:.0f}h")
    else:
        print("✗ people.yaml not found")

    # Heuristics
    heuristics = load_heuristics(edpa_root)
    if heuristics:
        print("✓ heuristics loaded")
    else:
        print("✗ heuristics not found (will use defaults)")

    # Iterations from edpa.yaml
    edpa_cfg_path = edpa_root / "config" / "edpa.yaml"
    if edpa_cfg_path.exists():
        edpa_cfg = load_yaml(edpa_cfg_path) or {}
        pi = edpa_cfg.get("pi", {})
        iterations = pi.get("iterations", [])
        if iterations:
            print(f"✓ {len(iterations)} iterations defined (PI: {pi.get('current', '?')})")
            for it in iterations:
                status = it.get("status", "?")
                marker = "→" if status == "active" else " "
                print(f"  {marker} {it.get('id', '?'):<16} {it.get('dates', ''):<20} [{status}]")
    else:
        print("✗ edpa.yaml not found")

    # Backlog
    backlog_dir = edpa_root / "backlog"
    if backlog_dir.exists():
        story_count = len(list((backlog_dir / "stories").glob("*.yaml"))) if (backlog_dir / "stories").exists() else 0
        feature_count = len(list((backlog_dir / "features").glob("*.yaml"))) if (backlog_dir / "features").exists() else 0
        print(f"✓ backlog — {feature_count} features, {story_count} stories")
    else:
        print("✗ backlog/ not found")

    # Reports
    reports_dir = edpa_root / "reports"
    if reports_dir.exists():
        report_dirs = [d for d in reports_dir.iterdir() if d.is_dir()]
        if report_dirs:
            print(f"✓ {len(report_dirs)} iteration report(s)")
        else:
            print("  reports/ empty (no iterations closed yet)")

    print()


def write_snapshot(edpa_root, iteration_id, engine_output, capacity):
    """Write frozen snapshot to .edpa/snapshots/ with revision tracking."""
    snapshots_dir = edpa_root / "snapshots"
    snapshots_dir.mkdir(parents=True, exist_ok=True)

    # Determine revision number
    base = snapshots_dir / f"{iteration_id}.json"
    if base.exists():
        # Find next revision
        rev = 2
        while (snapshots_dir / f"{iteration_id}_rev{rev}.json").exists():
            rev += 1
        snapshot_path = snapshots_dir / f"{iteration_id}_rev{rev}.json"
        print(f"Snapshot revision: {snapshot_path.name} (original exists)")
    else:
        snapshot_path = base

    snapshot = {
        "snapshot_version": VERSION,
        "iteration": iteration_id,
        "generated_at": engine_output["computed_at"],
        "frozen": True,
        "methodology": engine_output["methodology"],
        "mode": engine_output["mode"],
        "capacity_registry": {
            "people": capacity.get("people", []),
            "teams": capacity.get("teams", []),
        },
        "derived_reports": [
            {
                "person": r["id"],
                "name": r["name"],
                "role": r["role"],
                "capacity": r["capacity"],
                "total_derived": r["total_derived"],
                "items_count": len(r["items"]),
                "invariant_ok": r["invariant_ok"],
            }
            for r in engine_output["people"]
        ],
        "items": [],
        "invariants": {
            "all_passed": engine_output["all_invariants_passed"],
        },
        "signature_status": "pending",
    }

    # Collect all items with their contributors
    for person in engine_output["people"]:
        for item in person["items"]:
            snapshot["items"].append({
                "id": item["id"],
                "level": item["level"],
                "job_size": item["js"],
                "contributor": person["id"],
                "cw": item["cw"],
                "score": item["score"],
                "ratio": item["ratio"],
                "hours": item["hours"],
            })

    with open(snapshot_path, "w") as f:
        json.dump(snapshot, f, indent=2, ensure_ascii=False)
    print(f"Snapshot frozen: {snapshot_path}")


def write_excel(edpa_root, iteration_id, results, capacity):
    """Write summary.xlsx and item-costs.xlsx using openpyxl."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("Excel export skipped (install openpyxl for XLSX output)")
        return

    report_dir = edpa_root / "reports" / f"iteration-{iteration_id}"
    report_dir.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="2D2D2D", end_color="2D2D2D", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # --- summary.xlsx (per-person) ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    project_name = capacity.get("project", {}).get("name", "")
    ws.append([f"EDPA {VERSION} — {iteration_id}"])
    ws.merge_cells("A1:G1")
    ws["A1"].font = Font(bold=True, size=14)
    if project_name:
        ws.append([f"Project: {project_name}"])
        ws.merge_cells("A2:G2")
    ws.append([])

    headers = ["Person", "Role", "FTE", "Capacity (h)", "Derived (h)", "Items", "OK"]
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        fte = 0
        for p in capacity.get("people", []):
            if p.get("id") == r["id"]:
                fte = p.get("fte", 0)
                break
        row = [r["name"], r["role"], fte, r["capacity"],
               r["total_derived"], len(r["items"]),
               "✓" if r["invariant_ok"] else "✗"]
        ws.append(row)
        for col in range(1, len(row) + 1):
            ws.cell(row=ws.max_row, column=col).border = thin_border

    # Totals row
    total_cap = sum(r["capacity"] for r in results)
    total_derived = sum(r["total_derived"] for r in results)
    total_items = sum(len(r["items"]) for r in results)
    ws.append(["TOTAL", "", "", total_cap, total_derived, total_items, ""])
    for col in range(1, 8):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = header_font
        cell.border = thin_border

    # Auto-width (skip merged cells)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    summary_path = report_dir / "summary.xlsx"
    wb.save(summary_path)
    print(f"Excel: {summary_path}")

    # --- item-costs.xlsx (per-item) ---
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "Item Costs"

    ws2.append([f"EDPA {VERSION} — {iteration_id} — Per-Item Allocation"])
    ws2.merge_cells("A1:H1")
    ws2["A1"].font = Font(bold=True, size=14)
    ws2.append([])

    headers2 = ["Item", "Level", "JS", "Person", "CW", "Score", "Ratio", "Hours"]
    ws2.append(headers2)
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=ws2.max_row, column=col)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")

    for r in results:
        for item in r["items"]:
            row = [item["id"], item["level"], item["js"],
                   r["name"], item["cw"], round(item["score"], 2),
                   f"{item['ratio']:.1%}", round(item["hours"], 2)]
            ws2.append(row)
            for col in range(1, len(row) + 1):
                ws2.cell(row=ws2.max_row, column=col).border = thin_border

    for col_idx in range(1, ws2.max_column + 1):
        max_len = 0
        for row_idx in range(1, ws2.max_row + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 30)

    items_path = report_dir / "item-costs.xlsx"
    wb2.save(items_path)
    print(f"Excel: {items_path}")


def main():
    parser = argparse.ArgumentParser(
        description=f"EDPA {VERSION} — Evidence-Driven Proportional Allocation Engine",
        epilog="Run with --demo to see a worked example, or --edpa-root to read from .edpa/ filesystem."
    )
    parser.add_argument("--edpa-root", help="Path to .edpa/ directory (reads backlog, config, heuristics)")
    parser.add_argument("--iteration", help="Iteration ID (e.g., PI-2026-1.3)")
    parser.add_argument("--mode", choices=["simple", "full"], default="simple",
                        help="Calculation mode (default: simple)")
    parser.add_argument("--capacity", help="Path to capacity.yaml (legacy mode)")
    parser.add_argument("--heuristics", help="Path to cw_heuristics.yaml (legacy mode)")
    parser.add_argument("--output", help="Output path for edpa_results.json")
    parser.add_argument("--version", action="version", version=f"EDPA {VERSION}")
    parser.add_argument("--status", action="store_true",
                        help="Show EDPA setup status and exit")
    parser.add_argument("--demo", action="store_true",
                        help="Run with built-in sample data")
    args = parser.parse_args()

    if args.status:
        show_status(Path(args.edpa_root) if args.edpa_root else Path(".edpa"))
        sys.exit(0)

    if args.demo:
        print("Running EDPA demo with sample data...\n")
        capacity, heuristics, items = generate_demo_data()
        iteration_id = "DEMO-1.1"
    elif args.edpa_root:
        # Filesystem-first mode: read everything from .edpa/
        edpa_root = Path(args.edpa_root)
        if not edpa_root.exists():
            parser.error(f".edpa/ directory not found at {edpa_root}")

        capacity = load_yaml(edpa_root / "config" / "people.yaml")
        heuristics = load_heuristics(edpa_root)
        iteration_id = args.iteration

        items, manual_cw = load_backlog_items(edpa_root, iteration_id)
        print(f"Loaded {len(items)} items from {edpa_root}/backlog/")
        if iteration_id:
            print(f"Filtered to iteration: {iteration_id}")
        if manual_cw:
            print(f"Manual CW overrides: {len(manual_cw)}")
    else:
        # Legacy mode: explicit file paths
        if not args.capacity or not args.heuristics or not args.iteration:
            parser.error("--edpa-root or (--iteration + --capacity + --heuristics) required (or --demo)")

        capacity = load_yaml(args.capacity)
        heuristics = load_yaml(args.heuristics)
        iteration_id = args.iteration
        items = []
        print(f"Legacy mode: loading from {args.capacity} and {args.heuristics}")
        print(f"NOTE: No items loaded. Use --edpa-root to read from .edpa/backlog/")

    # Resolve planning_factor from teams (team-level decision, not cadence)
    teams = capacity.get("teams", [])
    if teams:
        planning_factor = teams[0].get("planning_factor", 0.8)
    else:
        planning_factor = 0.8

    results = run_edpa(capacity, heuristics, items, mode=args.mode)

    all_passed = all(r["invariant_ok"] for r in results if r["items"])
    team_total = sum(r["total_derived"] for r in results)

    output = {
        "iteration": iteration_id,
        "mode": args.mode,
        "computed_at": datetime.now(timezone.utc).isoformat(),
        "methodology": f"EDPA {VERSION}",
        "planning_factor": planning_factor,
        "people": results,
        "team_total": round(team_total, 2),
        "all_invariants_passed": all_passed,
    }

    # Write output
    if args.output:
        output_path = Path(args.output)
    elif args.edpa_root:
        output_path = Path(args.edpa_root) / "reports" / f"iteration-{iteration_id}" / "edpa_results.json"
    elif not args.demo:
        output_path = Path(f".edpa/reports/iteration-{iteration_id}/edpa_results.json")
    else:
        output_path = None

    if output_path:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        with open(output_path, "w") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)
        print(f"\nResults written to: {output_path}")

    # Write frozen snapshot
    if args.edpa_root and not args.demo:
        edpa_root = Path(args.edpa_root)
        write_snapshot(edpa_root, iteration_id, output, capacity)
        write_excel(edpa_root, iteration_id, results, capacity)

    print_summary(results, args.mode, iteration_id, planning_factor)

    if not all_passed:
        sys.exit(1)


if __name__ == "__main__":
    main()
